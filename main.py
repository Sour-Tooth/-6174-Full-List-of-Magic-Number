from openpyxl import Workbook
#from collections import Counter
import collections

magicNumber = 1234
cycles = 0
information = [0, 0, 0, 0, 0, 0, 0, 0]
stepInfo = []

wb = Workbook()
# use the worsheet
ws = wb.active

wb.create_sheet('Numbers Found During Steps')

ws.title = 'All magicNumbers Organized by Steps'

def removeDuplicatesFromString(string):
    result = ''
    for character in string:
        if not(character in result):
            result = result + character
    
    return result

def isAMagicNumber(number):
    string = str(number)

    if len(string) <= len(removeDuplicatesFromString(string)) + 2:
        return True
    else:
        return False

def takeAStep(numberToStep):
    lowToHi = ''.join(sorted(numberToStep))
    hiToLow = ''.join(sorted(numberToStep, reverse = True))
    result = str(int(hiToLow) - int(lowToHi))
    lowToHi.replace('0', '')

    result = str(int(hiToLow) - int(lowToHi))
    while len(result) != 4:
        result = result + '0'

    stepInfo.append(result)
    return result    

def stepUntilGoalReached(numberToStep):
    step = 0
    while numberToStep != '6174':
        numberToStep = takeAStep(numberToStep)
        step += 1
    
    print(f"Your magicNumber {magicNumber} took {step} steps to reach 6174.")
    information[step] += 1
    ws.cell(row = information[step], column = step + 1).value = f"{magicNumber}"

    return numberToStep

while magicNumber <= 9999:
    magicNumber = 1000 + cycles 

    if isAMagicNumber(magicNumber):
        stepUntilGoalReached(str(magicNumber))

    cycles += 1 

print(information)
#print(Counter(stepInfo).most_common())

test = collections.Counter(stepInfo)#.most_common()

print('The following section hilights all the numbers reaced after every step')
#for value in test:
#    print(f'{value} : {test[value]}')

for letter, count in test.most_common(): 
    print(f'{letter} : {count}')

# Save the file
wb.save("AllTheMagicNumbers.xlsx")